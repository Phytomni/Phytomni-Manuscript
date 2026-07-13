import re
from pathlib import Path

import nbformat
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIGURE_DIR = ROOT / "Extended Data Fig. 6"
NOTEBOOK = FIGURE_DIR / "extended_data_fig. 6abc.ipynb"
WORKBOOK = FIGURE_DIR / "PhytoBench-Knowledge-for_plot.xlsx"


def notebook_source() -> str:
    notebook = nbformat.read(NOTEBOOK, as_version=4)
    nbformat.validate(notebook)
    return "\n".join(
        cell.source for cell in notebook.cells if cell.cell_type == "code"
    )


def named_vector(source: str, variable: str) -> dict[str, str]:
    match = re.search(
        rf"{re.escape(variable)}\s*<-\s*c\((.*?)\n\)",
        source,
        flags=re.DOTALL,
    )
    assert match is not None, f"R vector {variable} is missing"
    return dict(re.findall(r'"([^"]+)"\s*=\s*"([^"]+)"', match.group(1)))


def test_model_maps_match_the_shipped_workbook_columns() -> None:
    source = notebook_source()
    pangu_model_map = named_vector(source, "pangu_model_map")
    comparison_model_map = named_vector(source, "comparison_model_map")

    assert pangu_model_map == {
        "Phyto-Chat-L0-Keyword": "Pangu-NLP-N2-128K (keyword)",
        "Phyto-Chat-L0-Semantic": "Pangu-NLP-N2-128K (semantic)",
        "Phyto-Chat-L0-Hybrid": "Pangu-NLP-N2-128K (hybrid)",
        "Phyto-Chat-L1-Keyword": "Phyto-Chatbot (keyword)",
        "Phyto-Chat-L1-Semantic": "Phyto-Chatbot (semantic)",
        "Phyto-Chat-L1-Hybrid": "Phyto-Chatbot (hybrid)",
        "Phyto-Reasoner-L0-Keyword": "Pangu-NLP-N2-Reasoner-128K (keyword)",
        "Phyto-Reasoner-L0-Semantic": "Pangu-NLP-N2-Reasoner-128K (semantic)",
        "Phyto-Reasoner-L0-Hybrid": "Pangu-NLP-N2-Reasoner-128K (hybrid)",
        "Phyto-Reasoner-L1-Keyword": "Phyto-Reasoner (keyword)",
        "Phyto-Reasoner-L1-Semantic": "Phyto-Reasoner (semantic)",
        "Phyto-Reasoner-L1-Hybrid": "Phyto-Reasoner (hybrid)",
    }
    assert comparison_model_map == {
        "Phyto-Chat-L1-Keyword": "Phyto-Chatbot (keyword)",
        "Phyto-Chat-L1-Semantic": "Phyto-Chatbot (semantic)",
        "Phyto-Chat-L1-Hybrid": "Phyto-Chatbot (hybrid)",
        "Phyto-Reasoner-L1-Keyword": "Phyto-Reasoner (keyword)",
        "Phyto-Reasoner-L1-Semantic": "Phyto-Reasoner (semantic)",
        "Phyto-Reasoner-L1-Hybrid": "Phyto-Reasoner (hybrid)",
        "DeepSeek-V3": "DeepSeek-V3",
        "DeepSeek-R1": "DeepSeek-R1",
        "Gemini-2.5-Pro": "Gemini-2.5-Pro",
        "Claude-4.1-Opus": "Claude-4.1-Opus",
        "GPT-5": "GPT-5",
        "Grok-3-Beta": "Grok-3-Beta",
        "O3": "O3",
    }

    workbook = load_workbook(WORKBOOK, read_only=True)
    try:
        workbook.active.reset_dimensions()
        headers = {cell.value for cell in next(workbook.active.iter_rows())}
    finally:
        workbook.close()

    source_models = set(pangu_model_map) | set(comparison_model_map)
    assert {f"{model}_check" for model in source_models} <= headers
    assert "transform_results(raw, pangu_model_map)" in source
    assert "transform_results(raw, comparison_model_map)" in source


def test_model_transform_fails_loudly_for_schema_drift() -> None:
    source = notebook_source()

    assert "missing_columns <- setdiff" in source
    assert "Missing expected model columns" in source
